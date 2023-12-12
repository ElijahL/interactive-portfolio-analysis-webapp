import openpyxl as pyxl
import os
import glob
import re
from pathlib import Path
from datetime import datetime

REPORT_FOLDER = r"\\dc-fa-task-01\PolyFAShare\PolyScenarioReports\Processed"
AGGREGATED_FOLDER = r"C:\MEG\visualization of db project\data\scenarios"
FILE_TEMPLATE_NAME = r"PolyScenarioReport_(0[1-9]|[12][0-9]|3[01])-(0[1-9]|1[1,2])-(19|20)\d{2}_\d{2}-\d{2}-\d{2}.xlsx"

STRAT_TO_FIND = "TWOU_CB"
SCENARIOS = ['EquityDown20Pct', 'EquityDown5Pct', 'EquityUp20Pct', 'EquityUp5Pct', 
                     'InterestRateDown25Bps', 'InterestRateDown75Bps', 'InterestRateUp25Bps', 'InterestRateUp75Bps', 
                     'CreditSpreadDown250Bps', 'CreditSpreadDown50Bps', 'CreditSpreadUp250Bps', 'CreditSpreadUp50Bps', 
                     'VolUp4Pct', 'VolDown4Pct', 'VolUp10Pct', 'VolDown10Pct', 
                     'CreditUp25EquityMinus10VolFlat', 'CreditUp25EquityMinus20VolFlat', 'CreditUp25EquityMinus30VolFlat', 'CreditUp50EquityMinus10VolFlat', 
                     'CreditUp50EquityMinus20VolFlat', 'CreditUp50EquityMinus30VolFlat', 'CreditUp100EquityMinus10VolFlat', 'CreditUp100EquityMinus20VolFlat', 
                     'CreditUp100EquityMinus30VolFlat', 'CreditUp25EquityMinus10VolUp25', 'CreditUp25EquityMinus20VolUp25', 'CreditUp25EquityMinus30VolUp25', 
                     'CreditUp50EquityMinus10VolUp25', 'CreditUp50EquityMinus20VolUp25', 'CreditUp50EquityMinus30VolUp25', 'CreditUp100EquityMinus10VolUp25', 
                     'CreditUp100EquityMinus20VolUp25', 'CreditUp100EquityMinus30VolUp25', 'CreditDown25EquityPlus10VolFlat', 'CreditDown25EquityPlus20VolFlat', 
                     'CreditDown25EquityPlus30VolFlat', 'CreditDown100EquityPlus10VolFlat', 'CreditDown100EquityPlus20VolFlat', 'CreditDown100EquityPlus30VolFlat', 
                     'CreditDown50EquityPlus10VolFlat', 'CreditDown50EquityPlus20VolFlat', 'CreditDown50EquityPlus30VolFlat', 
                     'Stagflation', 'FinancialCrisis', 'BullMarket', 'ModerateDeflation', 'InterestRateSpike', 
                     'RatesMinus100bp', 'PerfectStorm', 'PerfectStormVolUp', 'EquityCreditCrash', 
                     'CreditWiden', 'RatesUp', 'EquityCrash', 'BorrowTighten', 
                     'VolUp', 'VolDown', 'RatesDown', 'Basis', 'AEE1', 'AEE2', 'AEE3']
        

processed_dates = set()

rows_to_write = []
header = ",".join(["Date"] + SCENARIOS)
rows_to_write.append(header)

list_of_files = [os.path.join(REPORT_FOLDER, f) for f in os.listdir(REPORT_FOLDER) if re.search(FILE_TEMPLATE_NAME, f)]
#list_of_files = glob.glob(os.path.join(REPORT_FOLDER, FILE_TEMPLATE_NAME))
for fpath in list_of_files:
    report_date = Path(fpath).stem.split('_')[1]
    report_date = datetime.strptime(report_date, "%d-%m-%Y").strftime("%Y-%m-%d")
    if report_date in processed_dates:
        continue
    
    try:
        print(f"Reading {fpath}")
        wb = pyxl.load_workbook(fpath)
    except Exception as e:
        print(f"Couldn't read {fpath}. Error: {e}")
        continue


    try:
        cof0_ws = wb["RiskAEE"]
        col_names = {col[0].value:idx for idx, col in enumerate(cof0_ws.iter_cols(1, cof0_ws.max_column))}
        #print(col_names.keys())
        for row_cells in cof0_ws.iter_rows(min_row=2, max_row=cof0_ws.max_row):
            #if row_cells[col_names["Strategy Code"]].value == STRAT_TO_FIND:
            STRAT_TO_FIND = row_cells[col_names["Strategy Code"]].value
            if not STRAT_TO_FIND:
                continue

            range = [s for s in cof0_ws.merged_cells.ranges if row_cells[col_names["Strategy Code"]].coordinate in s]
            if not range:
                continue
            range = range[0]

            fpath = os.path.join(AGGREGATED_FOLDER, STRAT_TO_FIND + ".csv")
            with open(fpath, "a+") as f:
                if f.tell() == 0:
                    f.write(header + "\n")

                row_to_write = [report_date]
                for scen in SCENARIOS:
                    value = 0
                    for strategy_rows in cof0_ws.iter_rows(min_row=range.min_row, max_row=range.max_row):
                        value += strategy_rows[col_names[scen]].value or 0
                    row_to_write.append(str(value))
                
                f.write(",".join(row_to_write) + "\n")
            
            
            processed_dates.add(report_date)
    except Exception as e:
        print(f"Error: {e}")
        continue



print("Saved to file")

    



# fpath = os.path.join(REPORT_FOLDER, "PolyScenarioReport_09-10-2023_01-00-18.xlsx")
# wb = pyxl.load_workbook(fpath)
# strategies_sheet = wb["Strategies"]
# strategy_list = []
# for row in strategies_sheet.iter_rows(min_row=2, max_row=strategies_sheet.max_row, values_only=True):
#     strategy_name = row[0]
#     #scenario_fpath = os.path.join(AGGREGATED_FOLDER, strategy_name + ".xlsx")
#     #print(strategy_name)
#     strategy_list.append(strategy_name)

# for strategy in strategy_list[0:1]:
#     strategy_fpath = os.path.join(AGGREGATED_FOLDER, strategy_name + ".xlsx")
#     try:
#         wb = pyxl.load_workbook(fpath)
#     except Exception as e:
#         wb = pyxl.Workbook()


